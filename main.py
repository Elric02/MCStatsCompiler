import json
import os
import pandas as pd # type: ignore
import numpy as np # type: ignore
import configparser
import openpyxl # type: ignore
import datetime
import ftplib
import math
import warnings
import paramiko # type: ignore
import excel2img # type: ignore
import requests # type: ignore
import base64
import stat
import csv
import nbtlib




def list_sftp_directory(sftp, path="."):
    """List contents of directory and parent directory for debugging"""
    try:
        print(f"\nContents of current directory '{path}':")
        for entry in sftp.listdir_attr(path):
            print(f"{entry.filename:30} {'<DIR>' if stat.S_ISDIR(entry.st_mode) else '<FILE>'}")
        
        parent = os.path.dirname(path) if path != "/" else "/"
        print(f"\nContents of parent directory '{parent}':")
        for entry in sftp.listdir_attr(parent):
            print(f"{entry.filename:30} {'<DIR>' if stat.S_ISDIR(entry.st_mode) else '<FILE>'}")
    except Exception as e:
        print(f"Error listing directory: {e}")

def loadVanillaData(csvtoggle, csvpath, inputmode, ftpserver, ftppath):
    df = pd.DataFrame()
    if inputmode == "ftp" or inputmode == "sftp":
        if ftppath == "":
            ftppath_complete = "world/stats"
        else:
            ftppath_complete = ftppath + "/world/stats"
        if inputmode == "ftp":
            ftpserver.cwd(ftppath)
            with open("data/usercache/usercache.json", "wb") as file:
                ftpserver.retrbinary(f"RETR usercache.json", file.write)
            names = pd.DataFrame(json.load(open("data/usercache/usercache.json", "r")))
            # Go back to root
            ftpserver.cwd("../" * (len(ftpserver.pwd().split("/"))-1))
            # Get directories
            filenames = ftpserver.nlst(ftppath_complete)
            ftpserver.cwd(ftppath_complete)
        else:
            try:
                ftpserver.chdir(ftppath)
            except IOError:
                print(f"Failed to change to directory {ftppath}")
                list_sftp_directory(ftpserver)
                raise
            
            try:
                ftpserver.get("usercache.json", "data/usercache/usercache.json")
            except IOError:
                print("Failed to get usercache.json")
                list_sftp_directory(ftpserver)
                raise

            names = pd.DataFrame(json.load(open("data/usercache/usercache.json", "r")))
            
            try:
                current_path = ftpserver.getcwd()
                depth = len([x for x in current_path.split("/") if x]) if current_path != "/" else 0
                if depth > 0:
                    ftpserver.chdir("../" * depth)  # Return to root
                print(f"Trying to access {ftppath_complete}")
                filenames = ftpserver.listdir(ftppath_complete)
                ftpserver.chdir(ftppath_complete)
            except IOError:
                print(f"Failed to access {ftppath_complete}")
                list_sftp_directory(ftpserver)
                raise

        for filename in filenames:
            if filename[-1] == ".":
                continue
            filename = filename.split("/")[-1]
            print("Now processing", filename)
            # Download the file to process
            local_file = "data/stats"+filename
            with open(local_file, "wb") as file:
                if inputmode == "ftp":
                    ftpserver.retrbinary(f"RETR {filename}", file.write)
                else:
                    ftpserver.get(filename, local_file)
            with open(local_file, "r") as file:
                data = json.load(file)
            os.remove(local_file)
            
            # Import the JSON to a Pandas DF
            temp_df = pd.json_normalize(data, meta_prefix=True)
            temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
            temp_df = temp_df.transpose().iloc[1:].rename({0: temp_name.iloc[0]}, axis=1)
            # Split the index (stats.blabla.blabla) into 3 indexes (stats, blabla, blabla)
            temp_df.index = temp_df.index.str.split('.', expand=True)
            # If a stat name has a dot in it, remove the part after the dot
            if len(temp_df.index.levshape) > 3:
                temp_df.index = temp_df.index.droplevel(3)
                temp_df = temp_df.groupby(level=[0,1,2]).sum()
            #print(temp_df)
            #temp_df.to_csv('temp.csv')
            if df.empty:
                df = temp_df
            else:
                df = df.join(temp_df, how="outer")
        
        # Go back to root
        if inputmode == "ftp":
            ftpserver.cwd("../" * (len(ftpserver.pwd().split("/"))-1))
        else:
            current_path = ftpserver.getcwd()
            depth = len([x for x in current_path.split("/") if x]) if current_path != "/" else 0
            if depth > 0:
                ftpserver.chdir("../" * depth)
    else:
        names_file = open('data/usercache/usercache.json', 'r')
        names = pd.DataFrame(json.load(names_file))
        for filename in os.listdir('data/stats'):
            if filename == ".gitignore":
                continue
            print("Now processing", filename)
            file = open('data/stats/' + filename)
            data = json.load(file)
            # Import the JSON to a Pandas DF
            temp_df = pd.json_normalize(data, meta_prefix=True)
            temp_name = names.loc[names['uuid'] == filename[:-5]]['name']
            temp_df = temp_df.transpose().iloc[1:].rename({0: temp_name.iloc[0]}, axis=1)
            # Split the index (stats.blabla.blabla) into 3 indexes (stats, blabla, blabla)
            temp_df.index = temp_df.index.str.split('.', expand=True)
            # If a stat name has a dot in it, remove the part after the dot
            if len(temp_df.index.levshape) > 3:
                temp_df.index = temp_df.index.droplevel(3)
                temp_df = temp_df.groupby(level=[0,1,2]).sum()
            #print(temp_df)
            #temp_df.to_csv('temp.csv')
            if df.empty:
                df = temp_df
            else:
                df = df.join(temp_df, how="outer")
    
    # Replace missing values by 0 (the stat has simply not been initialized because the associated action was not performed)
    df = df.fillna(0)
    if csvtoggle == "true":
        df.to_csv(csvpath)
    return df


def loadCobblemonData(csvtoggle, csvpath, inputmode, ftpserver, ftppath):

    # DataFrames
    df_total = pd.DataFrame()
    df_shiny = pd.DataFrame()
    player_count = {}

    # Create local directory
    os.makedirs("data/pokedex", exist_ok=True)
    os.makedirs("data/usercache", exist_ok=True)

    # Absolute path to the pokedex's data
    ftppath_complete = "Minecraft/world/pokedex" if ftppath == "root" else ftppath + "/world/pokedex"

    # Get usercache.json
    local_usercache = "data/usercache/usercache.json"
    if inputmode == "ftp":
        ftpserver.cwd(ftppath)
        with open(local_usercache, "wb") as f:
            ftpserver.retrbinary("RETR usercache.json", f.write)
        names = pd.DataFrame(json.load(open(local_usercache, "r")))
        root_dirs = ftpserver.nlst(ftppath_complete)
    elif inputmode == "sftp":
        remote_usercache = "Minecraft/usercache.json"
        with ftpserver.open(remote_usercache, "r") as rf:
            data = rf.read()
        with open(local_usercache, "wb") as lf:
            lf.write(data)
        names = pd.DataFrame(json.load(open(local_usercache, "r")))
        root_dirs = [
            f.filename for f in ftpserver.listdir_attr(ftppath_complete)
            if stat.S_ISDIR(f.st_mode)
        ]

    # Go to folders (62, 8d…)
    for dirname in root_dirs:
        prefix_dir = f"{ftppath_complete}/{dirname}" if inputmode == "sftp" else dirname

        # Listing NBT files
        if inputmode == "ftp":
            ftpserver.cwd(prefix_dir)
            nbt_files = ftpserver.nlst()
            ftpserver.cwd("..")
        else:
            nbt_files = [
                f.filename for f in ftpserver.listdir_attr(prefix_dir)
                if not stat.S_ISDIR(f.st_mode)
            ]

        for filename in nbt_files:
            if not filename.endswith(".nbt"):
                continue

            print("Now processing", filename)
            uuid = filename[:-4]
            remote_file = f"{prefix_dir}/{filename}"
            local_file = f"data/pokedex/{filename}"

            # Download
            if inputmode == "ftp":
                with open(local_file, "wb") as f:
                    ftpserver.retrbinary(f"RETR {remote_file}", f.write)
            else:
                with ftpserver.open(remote_file, "r") as rf:
                    data = rf.read()
                with open(local_file, "wb") as lf:
                    lf.write(data)

            # Charge NBTlib
            try:
                nbt = nbtlib.load(local_file)
            except Exception as e:
                print("Error NBT :", e)
                continue

            if "speciesRecords" not in nbt:
                print("No speciesRecords in", filename)
                continue

            species_records = nbt["speciesRecords"]

            # Create dict for this player
            parsed_total = {}
            parsed_shiny = {}

            for species, content in species_records.items():
                species = str(species).split(":")[-1]
                form_records = content["formRecords"]

                for form_name, form_data in form_records.items():
                    knowledge = str(form_data["knowledge"])  # ENCOUNTERED/CAUGHT
                    seen_shiny = [str(x).lower() for x in form_data["seenShinyStates"]]

                    is_caught = int(knowledge == "CAUGHT")  # 1 if caught
                    is_shiny = int("shiny" in seen_shiny)    # 1 if shiny

                    index = f"{species}.{form_name}"
                    parsed_total[index] = is_caught
                    parsed_shiny[index] = is_shiny

            # Convert dict in df for the player
            temp_total_df = pd.DataFrame.from_dict(parsed_total, orient="index", columns=[uuid])
            temp_shiny_df = pd.DataFrame.from_dict(parsed_shiny, orient="index", columns=[uuid])

            # Find UUID → pseudo
            temp_name = names.loc[names['uuid'] == uuid, 'name']
            player_name = temp_name.iloc[0] if not temp_name.empty else uuid

            # handling duplicates
            player_count[player_name] = player_count.get(player_name, 0) + 1
            if player_count[player_name] > 1:
                player_name = f"{player_name}_{player_count[player_name]}"

            temp_total_df.columns = [player_name]
            temp_shiny_df.columns = [player_name]

            # add to final df
            df_total = temp_total_df if df_total.empty else df_total.join(temp_total_df, how="outer")
            df_shiny = temp_shiny_df if df_shiny.empty else df_shiny.join(temp_shiny_df, how="outer")

    # Replace NaN
    df_total = df_total.fillna(0)
    df_shiny = df_shiny.fillna(0)

    # Export CSV 
    if csvtoggle == "true":
        df_total.to_csv(csvpath.replace(".csv", "_total.csv"))
        df_shiny.to_csv(csvpath.replace(".csv", "_shiny.csv"))

    return df_total, df_shiny


def getVanillaLeaderboard(df, cat, subcat):
    row = df.loc['stats'].loc[cat].loc[subcat].sort_values()
    print("Leaderboard of", cat, subcat, ":")
    print(row)

def getVanillaBestAndWorst(df, username, cleaning, cleaningvalue):
    if username == "null" or not username:
        print("Erreur: Aucun nom d'utilisateur spécifié dans la configuration")
        return
        
    if username not in df.columns:
        print(f"Erreur: L'utilisateur '{username}' n'existe pas dans les données")
        print("Utilisateurs disponibles:", ", ".join(df.columns))
        return
        
    nb_players = df.shape[1]
    if cleaning == "true":
        before_value = df.shape[0]
        df['zero_count'] = df.apply(lambda row: (row == 0).sum(), axis=1)
        df.drop(df[df['zero_count'] > (nb_players-int(cleaningvalue))].index, inplace=True)
        df = df.drop('zero_count', axis=1)
        print(before_value - df.shape[0], "rows dropped out of", before_value, "because of cleaning.")
    ranks = df.rank(axis=1, method='min', ascending=False)
    ranks['non_zero_values'] = df.apply(lambda row: nb_players - (row == 0).sum(), axis=1)
    ranks['value'] = df[username]
    output = ranks[[username, 'value', 'non_zero_values']].sort_values(username, ascending=False).rename(columns={username:"rank_"+username, "value":"value_"+username})
    print(output) # add .to_string() for the whole output

def most_pokemons_leaderboard(df, config, type):
    # Load the Excel file
    file_path = "output.xlsx"
    wb = openpyxl.load_workbook(file_path)
    print("DEBUG player_sum:")
    print(player_sum.head(20))

    if type == "standard":
        sheet_name = "leaderboard2"
    elif type == "shiny":
        sheet_name = "leaderboard3"
    elif type == "legendary":
        sheet_name = "leaderboard4"
    ws = wb[sheet_name]
    i = 0
    ExcelRows = int(config['COBBLEMONLEADERBOARDS']['ExcelRows'])
    ExcelCols = int(config['COBBLEMONLEADERBOARDS']['ExcelColumns'])
    for index, row in df[0:ExcelRows*ExcelCols].iterrows():
        ws.cell(row=(i%ExcelRows)+3, column=2+math.floor(i/ExcelRows)*3, value=str(i+1)+".")
        ws.cell(row=(i%ExcelRows)+3, column=3+math.floor(i/ExcelRows)*3, value=index)
        ws.cell(row=(i%ExcelRows)+3, column=4+math.floor(i/ExcelRows)*3, value=row[0])
        i += 1
    now = datetime.datetime.now()
    ws.cell(row=ExcelRows+3, column=2, value=now.strftime(config['COBBLEMONLEADERBOARDS']['LastUpdated']))
    ws.cell(row=ExcelRows+4, column=2, value=config['COBBLEMONLEADERBOARDS']['Subtitle'])
    wb.save(file_path)


def export_excel_to_image(config):
    """Convert Excel sheets to images"""
    
    file_path = "output.xlsx"
    # Selection of the area to export
    selection = "A1:N15"
    
    try:
        if config['COBBLEMONLEADERBOARDS']['TotalEnable'] == "true":
        
            excel2img.export_img(
                file_path,
                "leaderboard2.png",
                "leaderboard2",
                selection
            )
            
        if config['COBBLEMONLEADERBOARDS']['ShinyEnable']== "true":
            excel2img.export_img(
                file_path,
                "leaderboard3.png",
                "leaderboard3",
                selection
            )
        
        if config['COBBLEMONLEADERBOARDS']['LegEnable'] == "true":
            excel2img.export_img(
                file_path,
                "leaderboard4.png",
                "leaderboard4",
                selection
            )

        
    except Exception as e:
        print("Erreur lors de l'exportation des images.")
        print(e)

def check_file_exists(api_url, headers):
    response = requests.get(api_url, headers=headers)
    return response.status_code == 200, response.json().get("sha") if response.status_code == 200 else None

def upload_image(api_url, headers, image_data):
    data = {
        "message": "Upload initial de l'image",
        "content": image_data,
        "branch": BRANCH
    }
    return requests.put(api_url, headers=headers, json=data)

def update_image(api_url, headers, image_data, sha):
    data = {
        "message": "Mise à jour de l'image",
        "content": image_data,
        "branch": BRANCH,
        "sha": sha
    }
    return requests.put(api_url, headers=headers, json=data)

# Read config
config = configparser.ConfigParser()
config.read('config.ini', encoding='utf8')

# Connect to FTP if activated
ftp_server = None
if config['INPUT']['Mode'] == "ftp":
    ftp_server = ftplib.FTP(config['INPUT']['Host'], open("username.txt", "r").read(), open("password.txt", "r").read())
    ftp_server.encoding = "utf-8"
if config['INPUT']['Mode'] == "sftp":
    transport = paramiko.Transport((config['INPUT']['Host'], int(config['INPUT']['Port'])))
    transport.connect(username=open("username.txt", "r").read().strip(), password=open("password.txt", "r").read().strip())
    ftp_server = paramiko.SFTPClient.from_transport(transport)

if config['VANILLALEADERBOARD']['Enable'] == "true":
    # Load the data
    print("LOADING VANILLA DATA")
    vanilla_df = loadVanillaData(config['VANILLALEADERBOARD']['CreateCSV'], config['VANILLALEADERBOARD']['CSVPath'], config['INPUT']['Mode'], ftp_server, config['INPUT']['FTPPath'])


if config['COBBLEMONLEADERBOARDS']['TotalEnable'] == "true" or config['COBBLEMONLEADERBOARDS']['ShinyEnable'] == "true" or config['COBBLEMONLEADERBOARDS']['LegEnable'] == "true":
    print("LOADING COBBLEMON DATA")
    if config['GLOBALMATRIX']['UseCSV'] == "false":
        df_total, df_shiny = loadCobblemonData(
            config['GLOBALMATRIX']['CreateCSV'],
            config['GLOBALMATRIX']['CSVPath'],
            config['INPUT']['Mode'],
            ftp_server,
            config['INPUT']['FTPPath']
        )
    else:
        cobblemon_df = pd.read_csv(config['GLOBALMATRIX']['CSVPath'], index_col=[0,1,2], skipinitialspace=True)

# Close the Connection
if config['INPUT']['Mode'] == "ftp":
    ftp_server.quit()
if config['INPUT']['Mode'] == "sftp":
    ftp_server.close()

# First leaderboard testing
if config['VANILLALEADERBOARD']['Enable'] == "true":
    getVanillaLeaderboard(vanilla_df, config['VANILLALEADERBOARD']['Category'], config['VANILLALEADERBOARD']['Subcategory'])

# First bestandworst testing
if config['BESTANDWORST']['Enable'] == "true":
    getVanillaBestAndWorst(vanilla_df, config['BESTANDWORST']['Username'], config['BESTANDWORST']['Cleaning'], config['BESTANDWORST']['CleaningValue'])

# Prepare the counting DF
# Colonnes à supprimer
cols_to_drop = ['caughtTimestamp', 'discoveredTimestamp', 'isShiny']

# Vérifie si le DataFrame a des colonnes MultiIndex
if isinstance(df_total.columns, pd.MultiIndex):
    # Sélectionne seulement les colonnes existantes dans n'importe quel niveau
    existing_cols = [col for col in df_total.columns.get_level_values(-1) if col in cols_to_drop]
    # Supprime les colonnes existantes
    count_df = df_total.drop(columns=existing_cols, level=-1, errors='ignore')
else:
    # DataFrame à colonnes simples
    count_df = df_total.drop(columns=[c for c in cols_to_drop if c in df_total.columns], errors='ignore')

pokemons_db = pd.read_csv('Pokemon.csv')
legendary_list = pokemons_db.loc[pokemons_db['Legendary'] == True]

# Total leaderboard feature
if config['COBBLEMONLEADERBOARDS']['TotalEnable'] == "true":
    # Somme par joueur → axis=0 car joueurs sont en colonnes
    player_sum = pd.DataFrame(df_total.sum(axis=0)).sort_values(by=0, ascending=False)
    player_sum['index'] = range(1, len(player_sum)+1)
    ignore_names = [name.strip() for name in config['COBBLEMONLEADERBOARDS']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    most_pokemons_leaderboard(player_sum, config, "standard")

# Shiny leaderboard
if config['COBBLEMONLEADERBOARDS']['ShinyEnable'] == "true":
    # Somme par joueur → axis=0
    player_sum = pd.DataFrame(df_shiny.sum(axis=0)).sort_values(by=0, ascending=False)
    player_sum['index'] = range(1, len(player_sum)+1)
    ignore_names = [name.strip() for name in config['COBBLEMONLEADERBOARDS']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    most_pokemons_leaderboard(player_sum, config, "shiny")
# Legendary leaderboard feature
if config['COBBLEMONLEADERBOARDS']['LegEnable'] == "true":
    legs = legendary_list['Cobblemon'].tolist()  # liste des légendaires
    
    # Filtrer seulement les lignes correspondantes aux légendaires
    leg_df = df_total.loc[df_total.index.isin(legs)]
    
    # Somme par joueur → axis=0
    player_sum = pd.DataFrame(leg_df.sum(axis=0)).sort_values(by=0, ascending=False)
    
    # Classement
    player_sum['index'] = range(1, len(player_sum)+1)
    
    # Supprimer les joueurs à ignorer
    ignore_names = [name.strip() for name in config['COBBLEMONLEADERBOARDS']['IgnoreNames'].split(",") if name.strip()]
    player_sum.drop(ignore_names, inplace=True, errors='ignore')
    
    # Générer le leaderboard
    most_pokemons_leaderboard(player_sum, config, "legendary")

# Export the Excel to images
export_excel_to_image(config)

if config['GIT']['UseGit'] == "true":
    GITHUB_USERNAME = config['GIT']['Username']
    REPO_NAME = config['GIT']['Repo']
    BRANCH = config['GIT']['Branch']
    GITHUB_TOKEN = config['GIT']['Token'].strip()
    
    # Vérification du token
    if not GITHUB_TOKEN or GITHUB_TOKEN.startswith('"') or GITHUB_TOKEN.endswith('"'):
        print(" Erreur : Le token GitHub est mal formaté. Assurez-vous qu'il n'y a pas de guillemets dans le fichier de configuration.")
        exit(1)

    try:
        if config['COBBLEMONLEADERBOARDS']['TotalEnable'] == "true":
            with open("leaderboard2.png", "rb") as img_file:
                img_data = base64.b64encode(img_file.read()).decode("utf-8")
            headers = {
                "Authorization": f"token {GITHUB_TOKEN}",
                "Accept": "application/vnd.github.v3+json"
            }
            file_name = os.path.basename("leaderboard2.png")
            api_url = f"https://api.github.com/repos/{GITHUB_USERNAME}/{REPO_NAME}/contents/{file_name}"
            
            exists, sha = check_file_exists(api_url, headers)
            if exists:
                print("Mise à jour de l'image existante...")
                response = update_image(api_url, headers, img_data, sha)
            else:
                print("Upload d'une nouvelle image...")
                response = upload_image(api_url, headers, img_data)
            if response.status_code in [200, 201]:
                print(f"Opération réussie : https://raw.githubusercontent.com/{GITHUB_USERNAME}/{REPO_NAME}/refs/heads/{BRANCH}/{file_name}")
            else:
                print(f"Erreur ({response.status_code}): {response.json()}")
    except Exception as e:
        print("Erreur lors de l'opération.")
        print(e)
    try:
        if config['COBBLEMONLEADERBOARDS']['ShinyEnable'] == "true":
            with open("leaderboard3.png", "rb") as img_file:
                img_data = base64.b64encode(img_file.read()).decode("utf-8")
            headers = {
                "Authorization": f"token {GITHUB_TOKEN}",
                "Accept": "application/vnd.github.v3+json"
            }
            file_name = os.path.basename("leaderboard3.png")
            api_url = f"https://api.github.com/repos/{GITHUB_USERNAME}/{REPO_NAME}/contents/{file_name}"
            exists, sha = check_file_exists(api_url, headers)
            if exists:
                print("Mise à jour de l'image existante...")
                response = update_image(api_url, headers, img_data, sha)
            else:
                print("Upload d'une nouvelle image...")
                response = upload_image(api_url, headers, img_data)
            if response.status_code in [200, 201]:
                print(f"Opération réussie : https://raw.githubusercontent.com/{GITHUB_USERNAME}/{REPO_NAME}/refs/heads/{BRANCH}/{file_name}")
            else:
                print(f"Erreur ({response.status_code}): {response.json()}")
    except Exception as e:
        print("Erreur lors de l'opération.")
        print(e)
    try:
        if config['COBBLEMONLEADERBOARDS']['LegEnable'] == "true":
            with open("leaderboard4.png", "rb") as img_file:
                img_data = base64.b64encode(img_file.read()).decode("utf-8")
            headers = {
                "Authorization": f"token {GITHUB_TOKEN}",
                "Accept": "application/vnd.github.v3+json"
            }
            file_name = os.path.basename("leaderboard4.png")
            api_url = f"https://api.github.com/repos/{GITHUB_USERNAME}/{REPO_NAME}/contents/{file_name}"
            exists, sha = check_file_exists(api_url, headers)
            if exists:
                print("Mise à jour de l'image existante...")
                response = update_image(api_url, headers, img_data, sha)
            else:
                print("Upload d'une nouvelle image...")
                response = upload_image(api_url, headers, img_data)
            if response.status_code in [200, 201]:
                print(f"Opération réussie : https://raw.githubusercontent.com/{GITHUB_USERNAME}/{REPO_NAME}/refs/heads/{BRANCH}/{file_name}")            
            else:
                print(f"Erreur ({response.status_code}): {response.json()}")
    except Exception as e:
        print("Erreur lors de l'opération.")
        print(e)
    
 

print("Done!")